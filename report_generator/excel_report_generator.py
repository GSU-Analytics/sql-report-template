# excel_report_generator.py

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font

class ExcelReportGenerator:
    def __init__(self, results, intro_text):
        """
        Initializes the Excel report generator.

        Args:
            results (dict): A dictionary where each key is a sheet name (derived from a SQL file name)
                            and each value is a dictionary mapping query titles to DataFrames.
            intro_text (list of str): A list of text lines to be added to the introduction sheet.
        """
        self.results = results
        self.intro_text = intro_text
        self.wb = Workbook()
        # Remove the default sheet if it exists.
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        # Global counter for table names to ensure uniqueness across the workbook.
        self.global_table_counter = 1

    def create_introduction_sheet(self):
        """
        Creates the Introduction worksheet and populates it with the intro text.
        """
        ws = self.wb.create_sheet("Introduction", 0)
        # Title for the introduction section.
        ws.append(["Introduction"])
        ws["A1"].font = Font(bold=True, size=14)
        
        # Append each line of intro text.
        for line in self.intro_text:
            ws.append([line])
        
        # Optionally, adjust the width of the first column.
        ws.column_dimensions['A'].width = 100

    def append_df_as_table(self, ws, df, table_title, start_row):
        """
        Appends a Pandas DataFrame to a worksheet as a formatted table.

        Args:
            ws (Worksheet): The worksheet object.
            df (DataFrame): The Pandas DataFrame to add.
            table_title (str): Title to display above the table.
            start_row (int): The starting row number in the worksheet.

        Returns:
            int: The row number after the inserted table.
        """
        # Write the table title with formatting.
        ws.cell(row=start_row, column=1, value=table_title).font = Font(bold=True, size=12)
        start_row += 1
        initial_data_row = start_row

        # Append the DataFrame rows (header and data).
        for i, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=0):
            ws.append(row)
        rows_appended = i + 1
        end_row = start_row + rows_appended - 1
        start_col = 1
        end_col = df.shape[1]

        # Create an Excel table with a unique name.
        table_ref = f"{get_column_letter(start_col)}{initial_data_row}:{get_column_letter(end_col)}{end_row}"
        table = Table(displayName=f"Table{self.global_table_counter}", ref=table_ref)
        self.global_table_counter += 1

        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        # Format any column whose name contains "Rate" as a percentage.
        rate_cols = [col for col in df.columns if "Rate" in col]
        for col_idx, col_name in enumerate(df.columns, start=1):
            if col_name in rate_cols:
                for row_idx in range(initial_data_row + 1, end_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = '0%'

        # Return the next available row (with a couple of blank rows added).
        return end_row + 2

    def create_results_sheets(self):
        """
        Creates a separate worksheet for each SQL file (based on the file name) and
        populates it with all tables derived from that file’s queries.
        Assumes that self.results is a dictionary with sheet names as keys and values
        that are dictionaries mapping query titles to DataFrames.
        """
        # Iterate over each file's results.
        for sheet_name, queries_dict in self.results.items():
            ws = self.wb.create_sheet(sheet_name)
            current_row = 1

            # Write each query result (table) from the file into the sheet.
            for table_title, df in queries_dict.items():
                current_row = self.append_df_as_table(ws, df, table_title, current_row)

            # Auto-adjust column widths.
            for col in ws.columns:
                max_length = max((len(str(cell.value)) if cell.value else 0 for cell in col), default=0)
                adjusted_width = max_length + 2
                ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    def generate_workbook(self, output_path):
        """
        Creates the workbook with the Introduction and all results sheets and saves it.

        Args:
            output_path (str): The file path where the workbook will be saved.
        """
        self.create_introduction_sheet()
        self.create_results_sheets()
        self.wb.save(output_path)
        print(f"Workbook successfully saved to {output_path}")
